import numpy as np
from datetime import time, timedelta
from openpyxl import load_workbook, Workbook
from pprint import pprint
from re import findall

DAYS_OF_WEEK = {
    "Monday": 0,
    "Tuesday": 1,
    "Wednesday": 2,
    "Thursday": 3,
    "Friday": 4,
    "Saturday": 5,
}

batches = {}
batch_years = {
    1: set(),
    2: set(),
    3: set(),
}
batch_tt = {}
year_timing = {}
lab_hours = []
lab_timing = {}
deanery_subjects = {}


def read_excel_subjects():
    """Read the excel sheet with hours requirements"""
    subject_wb = load_workbook("SubjectRequirements.xlsx")
    for sheet in subject_wb.sheetnames:
        active_ws = subject_wb[sheet]
        year = 1
        batch = ""
        pos = 1
        deanery_subjects[sheet] = []

        for row in active_ws.iter_rows(2):
            pos += 1
            year = active_ws[f"A{pos}"].value or year
            batch = active_ws[f"B{pos}"].value or batch
            course = active_ws[f"C{pos}"].value
            course_type = active_ws[f"D{pos}"].value
            hours = active_ws[f"E{pos}"].value
            batch_count = active_ws[f"F{pos}"].value
            alt_lab = active_ws[f"G{pos}"].value
            batch_years[year].add(batch)
            deanery_subjects[sheet].append(course)
            if batch not in batches:
                batches.update({batch: {"year": year}})
            batches[batch][course] = {
                "course_type": course_type,
                "hours": hours,
                "batch_count": batch_count,
                "alt_lab": alt_lab,
            }
            col_count = len(year_timing[year])
            row_count = 6
            batch_tt[batch] = [["" for _ in range(col_count)] for _ in range(row_count)]
            # batch_tt[batch].insert(0, year_timing[year])


def read_course_req_timing_sheet():
    """Read the sheet having general requirements"""
    course_wb = load_workbook("CourseRequirements.xlsx")

    # * Find the timings of the course
    timing_sheet = course_wb["Timings"]
    pos = 1
    for row in timing_sheet.iter_rows(min_col=2):
        for cell in row:
            if type(cell.value) is str:
                start_str, stop_str = cell.value.split("-")
                start_time = time.fromisoformat(f"{start_str.zfill(5)}:00")
                stop_time = time.fromisoformat(f"{stop_str.zfill(5)}:00")
                duration = timedelta(
                    minutes=(stop_time.hour * 60 + stop_time.minute)
                    - (start_time.hour * 60 + start_time.minute)
                )
                if pos not in year_timing:
                    year_timing[pos] = []
                year_timing[pos].append((start_time, stop_time, duration))
                for pos_lab, item in enumerate(lab_hours):
                    if start_time == item[0]:
                        break
                    if start_time < item[0]:
                        lab_hours.insert(pos_lab, (start_time, stop_time, duration))
                        break
                else:
                    lab_hours.append((start_time, stop_time, duration))
        pos += 1
    # pprint(year_timing)
    # pprint(lab_hours)


def read_course_admin_sheet():
    """"""
    # * Find mandatory classes
    course_wb = load_workbook("CourseRequirements.xlsx")
    mandatory_sheet = course_wb["Admin"]
    pos = 2
    for row in mandatory_sheet.iter_rows(min_row=2):
        subject = mandatory_sheet[f"A{pos}"].value
        day = mandatory_sheet[f"B{pos}"].value
        hour = mandatory_sheet[f"C{pos}"].value
        start_str = hour.split("-")[0]
        start_time = time.fromisoformat(f"{start_str.zfill(5)}:00")
        years_str = mandatory_sheet[f"D{pos}"].value
        if type(years_str) is int:
            years = [years_str]
        elif type(years_str) is str:
            years = [int(item) for item in findall("\d", years_str)]  # type: ignore
        for year in years:
            col = 0
            for timing in year_timing[year]:
                if timing[0] == start_time:
                    break
                col += 1
            row_value = DAYS_OF_WEEK[day]
            for batch in batch_years[year]:
                batch_tt[batch][row_value][col] = {
                    "subject": subject,
                    "constraint": 1,
                }
        pos += 1


def read_course_lab_sheet():
    course_wb = load_workbook("CourseRequirements.xlsx")
    lab_sheet = course_wb["LABS"]
    row_count = 6
    col_count = len(lab_hours)
    pos = 2
    for row in lab_sheet.iter_rows(min_row=2):
        subject = lab_sheet[f"A{pos}"].value
        amount = lab_sheet[f"B{pos}"].value
        for count in range(amount):
            if amount == 1:
                lab_name = f"{subject}"
            else:
                lab_name = f"{subject} {count + 1}"
            lab_timing[lab_name] = [
                ["" for _ in range(col_count)] for _ in range(row_count)
            ]
            # lab_timing[lab_name].insert(0, lab_hours)
        pos += 1




read_course_req_timing_sheet()
read_excel_subjects()
# pprint(batches)
# pprint(batch_years)
read_course_admin_sheet()
read_course_lab_sheet()
# pprint(batch_tt)


def write_to_workbook():
    tt_wb = Workbook()
    active_sheet = tt_wb.active
    if active_sheet is None:
        print("Error in getting sheet")
        return
    active_sheet.title = "1"
    year_sheet = {}
    for year in batch_years:
        if year == 1:
            year_sheet[year] = active_sheet
        else:
            year_sheet[year] = tt_wb.create_sheet(f"{year}")
    year_sheet["lab"] = tt_wb.create_sheet("LAB")

    for year in batch_years:
        for batch in batch_years[year]:
            tt = batch_tt[batch]
            # print(tt)
            timing_list = [batch]
            for timing in year_timing[year]:
                timing_list.append(f"{timing[0]}-{timing[1]}")
            year_sheet[year].append(timing_list)
            for row, day in zip(tt, DAYS_OF_WEEK):
                row_values = [day]
                for value in row:
                    if type(value) is dict:
                        row_values.append(value["subject"])
                    else:
                        row_values.append("")
                year_sheet[year].append(row_values)
            year_sheet[year].append([])

    for lab in lab_timing:
        tt = lab_timing[lab]
        # print(tt)
        timing_list = [lab]
        for timing in lab_hours:
            timing_list.append(f"{timing[0]}-{timing[1]}")
        year_sheet["lab"].append(timing_list)
        for row, day in zip(tt, DAYS_OF_WEEK):
            row_values = [day]
            for value in row:
                if type(value) is dict:
                    row_values.append(value["course"])
                else:
                    row_values.append("")
            year_sheet["lab"].append(row_values)
        year_sheet["lab"].append([])

    # def _create_time_sheet():
    #     row_value = 3
    #     for batch in batches:
    #         year1_sheet[f"A{row_value}"].value = batch
    #         row = year1_sheet.iter_rows(
    #             min_col=2, max_col=len(time_slots) + 1, min_row=row_value
    #         )
    #         # print(row)
    #         for col in row:
    #             for cell, time in zip(col, time_slots):
    #                 cell.value = time
    #         # for cell, time in zip(row, time_slots):
    #         #     print(cell, time)
    #         #     cell.value = time
    #         row_value += 1

    # _create_time_sheet()
    tt_wb.save("TestTT.xlsx")


write_to_workbook()
